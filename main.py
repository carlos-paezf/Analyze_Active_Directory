from openpyxl import Workbook
from pathlib import Path
from pyad import adgroup


class Search_Active_Directory():
    def __init__(self):
        self.parameters_path = Path("./parameters.txt")

        self.group_members = self.check_active_directory()


    def read_groups_from_file(self) -> list[str]:
        """
        This Python function reads and returns a list of groups from a file specified by the
        `parameters_path`.
        :return: A list of strings is being returned.
        """
        if not self.parameters_path.exists():
            raise FileNotFoundError("Archivo de parámetros no encontrado")
        
        with self.parameters_path.open('r', encoding="utf-8") as file:
            groups = [line.strip() for line in file if line.strip()]
            return groups
        
    
    def get_group_members(self, group_name: str) -> list[str]:
        """
        This function retrieves the members of a specified Active Directory group by its name.
        
        :param group_name: The `get_group_members` function takes a `group_name` parameter, which is a
        string representing the name of a group. This function attempts to retrieve the members of the
        specified group using the `adgroup.ADGroup.from_cn(group_name)` method and then returns a list
        of member common names (`
        :type group_name: str
        :return: A list of strings containing the common names (cn) of the members of the specified
        group. If an error occurs during the process, an empty list will be returned.
        """
        try:
            group = adgroup.ADGroup.from_cn(group_name)
            members = group.get_members()
            return [member.cn for member in members]
        except Exception as e:
            print(f"Se ha encontrado un error: {e}")
            return []
        
    
    def check_active_directory(self) -> dict[str, list[str]]:
        """
        The function `check_active_directory` reads group names from a file, retrieves members for each
        group, and returns a dictionary mapping group names to their members.
        :return: A dictionary is being returned, where the keys are group names (strings) and the values
        are lists of group members (strings).
        """
        group_list = self.read_groups_from_file()
        group_members = {}

        for group_name in group_list:
            print(f"Consultando miembros del grupo '{group_name}'")
            members = self.get_group_members(group_name)
            group_members[group_name] = members
        
        return group_members


class Generate_Excel():
    def __init__(self, group_members: dict[str, list[str]]):
        self.output_file = Path("./Grupos_y_Miembros_AD.xlsx")

        self.group_members = group_members

    
    def save_to_excel(self) -> None:
        """
        The `save_to_excel` function creates an Excel workbook with group and member data and saves it
        to a specified file location.
        """
        workbook = Workbook()
        
        sheet = workbook.active        
        sheet.title = "Grupos y Miembros AD"

        headers = list(group_members.keys())

        max_members = max(len(members) for members in group_members.values())

        for col, group_name in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=group_name)

            for row in range(max_members):
                for col, group_name in enumerate(headers, start=1):
                    members = sorted(group_members[group_name])

                    if row < len(members):
                        sheet.cell(row=row + 2, column=col, value=members[row])

        workbook.save(self.output_file)

        print(f"Datos guardados en '{self.output_file.absolute}'")




if __name__ == "__main__":
    group_members = Search_Active_Directory().group_members

    Generate_Excel(group_members)