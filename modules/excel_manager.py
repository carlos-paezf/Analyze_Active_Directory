from openpyxl import Workbook
from pathlib import Path
from pyad import adgroup

from measure_run_time import measure_run_time

class Excel_Manager():
    def __init__(self):
        self.output_file = Path("./Grupos_y_Miembros_AD.xlsx")

    
    @measure_run_time
    def save_to_excel(self, group_members: dict[str, list[str]]) -> tuple[bool, str]:
        """
        The `save_to_excel` function creates an Excel workbook with group and member data and saves it
        to a specified file location.
        """
        workbook = Workbook()
        
        sheet = workbook.active        
        sheet.title = "Grupos y Miembros AD"

        headers = list(group_members.keys())

        max_members = max(len(members) for members in group_members.values())

        for col, group_name in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=group_name)

            for row in range(max_members):
                for col, group_name in enumerate(headers, start=1):
                    members = sorted(group_members[group_name])

                    if row < len(members):
                        sheet.cell(row=row + 2, column=col, value=members[row])

        workbook.save(self.output_file)

        print(f"Datos guardados en '{self.output_file.name}'")

        return True, "El reporte Excel se ha generado correctamente"